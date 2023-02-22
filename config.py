import json
import time
from datetime import datetime
import openpyxl


class Config:
    def __init__(self):
        with open('cache.json', 'r', encoding='utf-8') as f:
            self.CACHE = json.load(f)
            self.starting_time = 0

    def check_cache(self, chat_id, parsed_message):
        if not chat_id in self.CACHE:
            self.CACHE[chat_id] = parsed_message
            self.CACHE[chat_id]["BOT_STATE"] = "start"
            self.CACHE[chat_id]["GAME_SCENE"] = 0
        else:
            self.CACHE[chat_id].update(parsed_message)

    def update_cache(self, chat_id, parsed_message, bot_state, game_scene=0):
        self.CACHE[chat_id] = parsed_message
        self.CACHE[chat_id]["BOT_STATE"] = bot_state
        self.CACHE[chat_id]["GAME_SCENE"] = game_scene
        self.save_cache()

    def reset_cache(self, chat_id):
        self.CACHE[chat_id]["BOT_STATE"] = "start"
        self.CACHE[chat_id]["GAME_SCENE"] = 0
        self.save_cache()

    def get_cached_data(self, chat_id):
        if chat_id in self.CACHE:
            message_type = self.CACHE[chat_id]['message_type']
            message_id = self.CACHE[chat_id]['message_id']
            data = self.CACHE[chat_id]['data']

    def find_states_by_chat_id(self, chat_id):
        if chat_id in self.CACHE:
            bot_state = self.CACHE[chat_id]["BOT_STATE"]
            game_scene = self.CACHE[chat_id]["GAME_SCENE"]
        else:
            bot_state = "start"
            game_scene = 0
        return bot_state, game_scene

    def save_cache(self):
        with open('cache.json', 'w', encoding='utf-8') as f:
            json.dump(self.CACHE, f, indent=4, ensure_ascii=False)

    def load_worksheet(self):
        wb = openpyxl.load_workbook('times.xlsx')
        return wb

    def start_timer(self):
        self.start_date = datetime.today()
        starting_time = datetime.today()
        print("starting time: ", starting_time)

        wb = self.load_worksheet()
        ws = wb['times']

        row = 2
        while starting_time.strftime('%d.%m.%Y') != ws['A' + str(row)].value.strftime('%d.%m.%Y'):
            row += 1

        if ws['B' + str(row)].value == None:
            ws['B' + str(row)] = starting_time.strftime('%H:%M')

        wb.save('times.xlsx')

        with open('timer.txt', 'r+') as f:
            timer_and_row = f.read()
            if timer_and_row == '':
                f.write(str(time.time()) + "\n" + str(row))

    # self.starting_timestamp = time.time()

    def stop_timer(self):
        wb = self.load_worksheet()
        ws = wb['times']
        # self.row = 2
        # while ws['B' + str(self.row)].value == None:
        # 	print(ws['B' + str(self.row)].value)
        # 	self.row += 1
        with open('timer.txt', 'r+') as f:
            timer_and_row = f.read()
            timer_and_row = timer_and_row.split("\n")
            timer = float(timer_and_row[0])
            row = timer_and_row[1]
            if timer_and_row != '':
                f.truncate(0)

        ws['C' + row] = datetime.today().strftime('%H:%M')
        total_time = time.time() - timer
        # total_time = time.time() - self.starting_timestamp
        if ws['D' + row].value == None:
            ws['D' + row] = total_time
        else:
            ws['D' + row] = ws['D' + row].value + total_time
        wb.save('times.xlsx')
        return total_time

    def get_times(self, date):
        wb = self.load_worksheet()
        ws = wb['times']
        row = 2
        # current_cell = ws['A' + str(row)].value
        # while date != current_cell.strftime('%d.%m.%Y'):
        # 	row += 1
        # 	current_cell = ws['A' + str(row)].value
        # 	print(row)
        while ws['A' + str(row)].value.strftime('%d.%m.%Y') != date:
            row += 1
            if ws['A' + str(row)].value == None:
                return "This date does not exist in the database"

        total_seconds = ws['D' + str(row)].value
        if total_seconds == None:
            return "You did not have any timer session on this date"
        if total_seconds == None:
            return "This date does not exist in the database"
        text = f"Time started: {ws['B' + str(row)].value}\n" \
               f"Time finished: {ws['C' + str(row)].value}\n" \
               f"Total time in seconds: {round(total_seconds)}\n" \
               f"Total time in minutes: {round(total_seconds / 60)}\n" \
               f"Total time in hours: {round(total_seconds / 3600)}\n"
        return text

